from __future__ import annotations

import hashlib
import json
from datetime import UTC, datetime
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from metro_scenario_studio.domain.schemas import (
    AuditEvent,
    ExecutionResult,
    OriginType,
    ScenarioExecution,
    ScenarioStatus,
)

if TYPE_CHECKING:
    from pathlib import Path

EXCEL_SCHEMA_VERSION = "mss_excel_v1"
EXCEL_REQUIRED_SHEETS = [
    "resumen",
    "configuracion",
    "variables_externas",
    "eventos",
    "meteorologia",
    "prediccion_agregada",
    "prediccion_estacion_linea",
    "explicacion",
    "trazabilidad",
    "metadata_json",
]
EXCEL_TEXT_CHUNK_SIZE = 30000


def export_execution_to_excel(
    execution: ScenarioExecution,
    result: ExecutionResult,
    output_path: Path,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    _write_rows(
        workbook,
        "resumen",
        [
            {
                "execution_id": execution.id,
                "estado": execution.status.value,
                "origen": execution.origin_type.value,
                "rango_inicio": execution.range_start.isoformat(),
                "rango_fin": execution.range_end.isoformat(),
                "total_predicho": _network_total(result),
                "total_real": _network_real_total(result),
                "real_disponible": execution.real_data_status,
                "modelo": execution.model_name,
                "variante": execution.model_variant,
                "dataset_version": execution.dataset_version,
                "fecha_exportacion": datetime.now(UTC).isoformat(),
                "warnings": "; ".join(execution.warnings),
                "resumen_narrativo": result.narrative_summary or "",
            }
        ],
    )
    _write_rows(
        workbook,
        "configuracion",
        [
            {
                "campo": "natural_language_comment",
                "valor_auto": "",
                "valor_final": execution.input.natural_language_comment or "",
                "origen": "user",
                "modificado": bool(execution.input.natural_language_comment),
                "usado_en_modelo": False,
                "comentario": execution.comment or "",
            },
            {
                "campo": "model_variant",
                "valor_auto": "strict_available",
                "valor_final": execution.model_variant,
                "origen": "system",
                "modificado": execution.model_variant != "strict_available",
                "usado_en_modelo": True,
                "comentario": "Variante usada para la ejecucion.",
            },
        ],
    )
    _write_rows(
        workbook,
        "variables_externas",
        [
            {
                "fecha": item.target_date.isoformat(),
                "tipo_variable": "calendar",
                "nombre": "day_of_week",
                "valor_auto": item.day_of_week,
                "valor_final": item.day_of_week,
                "fuente": item.source,
                "cobertura": "available",
                "modificado": item.modified,
                "usado_en_modelo": item.used_by_model,
            }
            for item in execution.input.calendar_final
        ],
    )
    _write_rows(
        workbook,
        "eventos",
        [
            {
                "event_id": item.event_id,
                "nombre": item.name,
                "fecha": item.target_date.isoformat(),
                "modo_fechas": item.date_mode,
                "fecha_inicio": item.start_date.isoformat() if item.start_date else "",
                "fecha_fin": item.end_date.isoformat() if item.end_date else "",
                "fechas_seleccionadas": ", ".join(value.isoformat() for value in item.selected_dates),
                "todo_el_dia": item.all_day,
                "hora_inicio": item.start_time or "",
                "hora_fin": item.end_time or "",
                "tipo": item.event_type.value,
                "impacto_esperado": item.impact_level.value,
                "estaciones_afectadas": ", ".join(item.affected_stations),
                "comentario": item.comment or "",
                "fuente_justificacion": item.source or "",
                "evento_origen": item.origin_event_id or "",
                "borrado": item.deleted,
                "accion": "modified" if item.modified else "automatic",
            }
            for item in execution.input.events_final
        ],
    )
    _write_rows(
        workbook,
        "meteorologia",
        [
            {
                "fecha": item.target_date.isoformat(),
                "lluvia": item.rain,
                "lluvia_intensa": item.heavy_rain,
                "temperatura_aprox": item.approx_temperature,
                "dia_caluroso": item.hot_day,
                "dia_frio": item.cold_day,
                "mal_tiempo": item.bad_weather,
                "temp_min": item.temp_min,
                "temp_media": item.temp_mean,
                "temp_max": item.temp_max,
                "precip_mm": item.precip_mm,
                "horas_lluvia": item.rain_hours,
                "viento": item.wind,
                "humedad": item.humidity,
                "weather_code": item.weather_code or "",
                "alerta": item.alert_level or "",
                "resumen_alerta": item.alert_summary or "",
            }
            for item in execution.input.weather_final
        ],
    )
    _write_rows(workbook, "prediccion_agregada", [row.model_dump(mode="json") for row in result.aggregates])
    _write_rows(workbook, "prediccion_estacion_linea", [row.model_dump(mode="json") for row in result.prediction_rows])
    _write_rows(workbook, "explicacion", [row.model_dump(mode="json") for row in result.explanations])
    _write_rows(workbook, "trazabilidad", [row.model_dump(mode="json") for row in result.audit_events])

    payload = {
        "execution": execution.model_dump(mode="json"),
        "result": result.model_dump(mode="json"),
    }
    payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    checksum = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()
    created_at = datetime.now(UTC).isoformat()
    payload_chunks = _chunk_text(payload_json)
    _write_rows(
        workbook,
        "metadata_json",
        [
            {
                "schema_version": EXCEL_SCHEMA_VERSION,
                "payload_json": chunk,
                "checksum": checksum,
                "source_app": "Metro Scenario Studio",
                "created_at": created_at,
                "chunk_index": index,
                "chunk_count": len(payload_chunks),
            }
            for index, chunk in enumerate(payload_chunks, start=1)
        ],
    )

    for worksheet in workbook.worksheets:
        _style_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def import_execution_from_excel(path: Path) -> tuple[ScenarioExecution, ExecutionResult]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    missing = set(EXCEL_REQUIRED_SHEETS).difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {sorted(missing)}")

    metadata_rows = _read_sheet(workbook["metadata_json"])
    if not metadata_rows:
        raise ValueError("Workbook metadata_json sheet is empty.")
    metadata = metadata_rows[0]
    if metadata.get("schema_version") != EXCEL_SCHEMA_VERSION:
        raise ValueError("Unsupported Excel schema version.")
    payload_json = _metadata_payload_json(metadata_rows)
    checksum = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()
    if checksum != metadata.get("checksum"):
        raise ValueError("Workbook metadata checksum does not match payload_json.")

    payload = json.loads(payload_json)
    original_execution = ScenarioExecution.model_validate(payload["execution"])
    result = ExecutionResult.model_validate(payload["result"])
    imported_execution = original_execution.model_copy(
        update={
            "id": f"imp_{checksum[:12]}",
            "status": ScenarioStatus.IMPORTADA,
            "origin_type": OriginType.EXCEL_IMPORT,
            "parent_execution_id": original_execution.id,
            "created_at": datetime.now(UTC),
            "updated_at": datetime.now(UTC),
        }
    )
    imported_result = result.model_copy(
        update={
            "execution": imported_execution,
            "audit_events": [
                *result.audit_events,
                AuditEvent(
                    execution_id=imported_execution.id,
                    action="import_excel",
                    summary=f"Imported from {path.name}",
                    payload={"source_path": str(path)},
                ),
            ],
        }
    )
    return imported_execution, imported_result


def _write_rows(workbook: Workbook, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    if rows:
        headers = list(rows[0].keys())
    else:
        headers = ["empty"]
        rows = [{"empty": ""}]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([_excel_value(row.get(header)) for header in headers])


def _read_sheet(worksheet) -> list[dict[str, Any]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [{headers[index]: value for index, value in enumerate(row)} for row in rows[1:]]


def _chunk_text(value: str) -> list[str]:
    return [value[index : index + EXCEL_TEXT_CHUNK_SIZE] for index in range(0, len(value), EXCEL_TEXT_CHUNK_SIZE)] or [
        ""
    ]


def _metadata_payload_json(metadata_rows: list[dict[str, Any]]) -> str:
    if "chunk_index" not in metadata_rows[0]:
        return str(metadata_rows[0]["payload_json"])
    ordered_rows = sorted(metadata_rows, key=lambda row: int(row.get("chunk_index") or 0))
    expected_count = int(metadata_rows[0].get("chunk_count") or len(ordered_rows))
    if len(ordered_rows) != expected_count:
        raise ValueError("Workbook metadata_json chunks are incomplete.")
    return "".join(str(row.get("payload_json") or "") for row in ordered_rows)


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _style_sheet(worksheet) -> None:
    fill = PatternFill("solid", fgColor="6B130C")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 80)


def _network_total(result: ExecutionResult) -> float:
    for aggregate in result.aggregates:
        if aggregate.level == "network" and aggregate.target_date is None:
            return aggregate.y_pred
    return sum(row.y_pred for row in result.prediction_rows)


def _network_real_total(result: ExecutionResult) -> float | None:
    for aggregate in result.aggregates:
        if aggregate.level == "network" and aggregate.target_date is None:
            return aggregate.y_real
    values = [row.y_real for row in result.prediction_rows if row.y_real is not None]
    return sum(values) if values else None
