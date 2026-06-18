from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from metro_scenario_studio.domain.schemas import (
    AggregateRow,
    ExecutionResult,
    ExplanationItem,
    OriginType,
    PredictionRow,
    ScenarioExecution,
    ScenarioInput,
    ScenarioStatus,
)
from metro_scenario_studio.services.excel_service import (
    EXCEL_REQUIRED_SHEETS,
    export_execution_to_excel,
    import_execution_from_excel,
)


def test_excel_export_contains_required_sheets_and_metadata_json(tmp_path) -> None:
    execution = _sample_execution()
    result = _sample_result(execution)
    output_path = tmp_path / "scenario.xlsx"

    export_execution_to_excel(execution, result, output_path)

    workbook = load_workbook(output_path, read_only=True)

    assert set(EXCEL_REQUIRED_SHEETS).issubset(set(workbook.sheetnames))
    metadata = workbook["metadata_json"]
    assert metadata["A1"].value == "schema_version"
    assert metadata["B1"].value == "payload_json"
    assert metadata["C1"].value == "checksum"
    assert metadata["B2"].value is not None


def test_excel_import_roundtrip_reconstructs_execution_as_imported(tmp_path) -> None:
    execution = _sample_execution()
    result = _sample_result(execution)
    output_path = tmp_path / "scenario.xlsx"

    export_execution_to_excel(execution, result, output_path)
    imported_execution, imported_result = import_execution_from_excel(output_path)

    assert imported_execution.status == ScenarioStatus.IMPORTADA
    assert imported_execution.origin_type == OriginType.EXCEL_IMPORT
    assert imported_execution.parent_execution_id == execution.id
    assert imported_execution.range_start == execution.range_start
    assert imported_result.prediction_rows[0].series_id == "ATZ"
    assert imported_result.aggregates[0].y_pred == result.aggregates[0].y_pred
    assert imported_result.narrative_summary == result.narrative_summary


def test_excel_import_roundtrip_handles_metadata_larger_than_excel_cell_limit(tmp_path) -> None:
    execution = _sample_execution()
    result = _sample_result(execution)
    result.prediction_rows = [
        result.prediction_rows[0].model_copy(
            update={
                "target_date": date(2026, 5, 20 + day),
                "series_id": f"ATZ_{day}_{station}",
                "estacion": f"Estacion con nombre largo {day}-{station}",
                "y_pred": float(100 + day + station),
            }
        )
        for day in range(6)
        for station in range(30)
    ]
    output_path = tmp_path / "large_scenario.xlsx"

    export_execution_to_excel(execution, result, output_path)
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    metadata_rows = list(workbook["metadata_json"].iter_rows(values_only=True))
    imported_execution, imported_result = import_execution_from_excel(output_path)

    assert len(metadata_rows) > 2
    assert imported_execution.status == ScenarioStatus.IMPORTADA
    assert len(imported_result.prediction_rows) == len(result.prediction_rows)


def _sample_execution() -> ScenarioExecution:
    return ScenarioExecution(
        id="scn_test",
        status=ScenarioStatus.WHAT_IF,
        real_data_status="datos reales disponibles",
        origin_type=OriginType.MANUAL,
        parent_execution_id=None,
        range_start=date(2026, 5, 20),
        range_end=date(2026, 5, 20),
        author="tester",
        comment="Partido y lluvia",
        model_name="tabular_hgbr",
        model_variant="forecastable_scenario",
        dataset_version="local-test",
        warnings=["missing_future_weather"],
        input=ScenarioInput(
            natural_language_comment="Partido y lluvia",
            calendar_final=[],
            events_final=[],
            weather_final=[],
            manual_overrides=[{"type": "weather", "field": "rain_expected"}],
            llm_detected_items=[],
            llm_accepted_items=[],
            llm_rejected_items=[],
        ),
    )


def _sample_result(execution: ScenarioExecution) -> ExecutionResult:
    row = PredictionRow(
        target_date=date(2026, 5, 20),
        linea="LINEA 1",
        estacion="Atarazanas",
        series_id="ATZ",
        station_abbrev="ATZ",
        network_order=1,
        y_pred=100.0,
        y_real=95.0,
        model_variant=execution.model_variant,
        horizon_days=1,
    )
    return ExecutionResult(
        execution=execution,
        prediction_rows=[row],
        aggregates=[
            AggregateRow(
                level="network",
                target_date=None,
                linea=None,
                estacion=None,
                y_pred=100.0,
                y_real=95.0,
                real_available=True,
                abs_error=5.0,
                pct_error=0.05263157894736842,
                delta_abs=None,
                delta_pct=None,
            )
        ],
        explanations=[
            ExplanationItem(
                section="variables_usadas",
                item_type="weather",
                label="Lluvia",
                description="Factor asociado usado por el modelo.",
                source="manual",
                used_by_model=True,
                confidence="medium",
                limitation=None,
            )
        ],
        narrative_summary="La demanda prevista sube de forma asociada al evento y la lluvia del escenario.",
        audit_events=[],
    )
