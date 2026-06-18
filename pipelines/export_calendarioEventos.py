from __future__ import annotations

import hashlib
import json
import logging
import os
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from requests_negotiate_sspi import HttpNegotiateAuth


BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / ".env"
OUT_DIR = BASE_DIR / "out"
LOGS_DIR = BASE_DIR / "logs"

load_dotenv(ENV_PATH)

# SharePoint upload dependency removed


ATOM_NS = "http://www.w3.org/2005/Atom"
M_NS = "http://schemas.microsoft.com/ado/2007/08/dataservices/metadata"


DEFAULT_EXCLUDE_COLUMNS = {
    "FirstUniqueAncestorSecurableObject",
    "RoleAssignments",
    "AttachmentFiles",
    "ContentType",
    "FieldValuesAsHtml",
    "FieldValuesAsText",
    "FieldValuesForEdit",
    "File",
    "Folder",
    "ParentList",
    "FileSystemObjectType",
    "GUID",
    # Campos base/técnicos habituales que no interesan en Excel
    # "Author",
    # "Editor",
    # "AuthorId",
    # "EditorId",
    "_UIVersionString",
    "OData__UIVersionString",
    "Attachments",
    "Edit",
    "LinkTitleNoMenu",
    "LinkTitle",
    "DocIcon",
    "ItemChildCount",
    "FolderChildCount",
    "AppAuthor",
    "AppEditor",
    "ContentTypeId",
}
UNSAFE_SELECT_TYPES = {
    "User",
    "UserMulti",
    "Lookup",
    "LookupMulti",
    "Computed",
    "ModStat",
    "ContentTypeId",
    "Attachments",
}


def parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "si", "sí"}


def split_csv(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def clean_excel_table_name(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in value)
    cleaned = cleaned.strip("_") or "TablaDatos"

    if cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"

    return cleaned[:255]


def clean_sheet_name(value: str) -> str:
    invalid = r"[]:*?/\\"
    cleaned = "".join("_" if ch in invalid else ch for ch in value).strip()
    return (cleaned or "Datos")[:31]


def parse_sp_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value

    if value is None:
        return False

    return str(value).strip().lower() in {"true", "1", "yes", "y", "sí", "si"}


@dataclass(frozen=True)
class Settings:
    source_site_url: str
    source_list_name: str
    output_name: str
    page_size: int

    dest_folder_url: str

    sheet_name: str
    columns_mode: str
    exclude_columns: list[str]
    drop_api_metadata: bool
    include_metadata_sheet: bool
    prevent_empty_upload: bool

    output_path: Path
    log_path: Path

    @classmethod
    def from_env(cls) -> "Settings":
        source_list_name = os.getenv("SOURCE_LIST_NAME", "").strip()
        output_name = os.getenv("OUTPUT_NAME", "").strip() or source_list_name

        if len(sys.argv) > 1 and sys.argv[1].endswith(".xlsx"):
            output_path = Path(sys.argv[1]).resolve()
        else:
            output_path = OUT_DIR / f"{output_name}.xlsx"
        log_path = LOGS_DIR / f"{output_name}.log"

        exclude_columns = split_csv(os.getenv("EXCLUDE_COLUMNS"))
        if not exclude_columns:
            exclude_columns = sorted(DEFAULT_EXCLUDE_COLUMNS)

        settings = cls(
            source_site_url=os.getenv("SOURCE_SITE_URL", "").rstrip("/"),
            source_list_name=source_list_name,
            output_name=output_name,
            page_size=int(os.getenv("PAGE_SIZE", "5000")),
            dest_folder_url=os.getenv("DEST_FOLDER_URL", "").rstrip("/"),
            sheet_name=clean_sheet_name(os.getenv("SHEET_NAME", "Datos")),
            columns_mode=os.getenv("COLUMNS_MODE", "auto").strip().lower(),
            exclude_columns=exclude_columns,
            drop_api_metadata=parse_bool(os.getenv("DROP_API_METADATA"), default=True),
            include_metadata_sheet=parse_bool(os.getenv("INCLUDE_METADATA_SHEET"), default=True),
            prevent_empty_upload=parse_bool(os.getenv("PREVENT_EMPTY_UPLOAD"), default=True),
            output_path=output_path,
            log_path=log_path,
        )

        settings.validate()
        return settings

    @property
    def dest_relative_file_url(self) -> str:
        return f"{self.dest_folder_url}/{self.output_name}.xlsx"

    @property
    def exclude_columns_lower(self) -> set[str]:
        return {col.lower() for col in self.exclude_columns}

    def validate(self) -> None:
        required = {
            "SOURCE_SITE_URL": self.source_site_url,
            "SOURCE_LIST_NAME": self.source_list_name,
            "OUTPUT_NAME": self.output_name,
        }

        missing = [key for key, value in required.items() if not value]
        if missing:
            raise ValueError(f"Faltan variables obligatorias: {', '.join(missing)}")

        if self.columns_mode not in {"auto", "all"}:
            raise ValueError("COLUMNS_MODE debe ser 'auto' o 'all'.")


def configure_logging(settings: Settings) -> None:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(settings.log_path, encoding="utf-8"),
        ],
        force=True,
    )


def make_session() -> requests.Session:
    session = requests.Session()
    session.auth = HttpNegotiateAuth()
    session.headers.update(
        {
            "Accept": "application/json;odata=verbose",
            "Content-Type": "application/json;odata=verbose",
            "User-Agent": "MetroMalaga-SharePointExport/1.0",
        }
    )
    return session


def local_name(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def convert_xml_value(text: str | None, edm_type: str | None, is_null: bool) -> Any:
    if is_null:
        return None

    if text is None:
        return ""

    try:
        if edm_type in {"Edm.Int16", "Edm.Int32", "Edm.Int64"}:
            return int(text)

        if edm_type in {"Edm.Double", "Edm.Decimal", "Edm.Single"}:
            return float(text)

        if edm_type == "Edm.Boolean":
            return text.lower() == "true"

        return text

    except Exception:
        return text


def parse_atom_feed(xml_text: str) -> tuple[list[dict[str, Any]], str | None]:
    root = ET.fromstring(xml_text)

    rows: list[dict[str, Any]] = []

    for entry in root.findall(f"{{{ATOM_NS}}}entry"):
        props = entry.find(f".//{{{M_NS}}}properties")
        if props is None:
            continue

        row: dict[str, Any] = {}

        for child in list(props):
            name = local_name(child.tag)
            edm_type = child.attrib.get(f"{{{M_NS}}}type")
            is_null = child.attrib.get(f"{{{M_NS}}}null") == "true"
            row[name] = convert_xml_value(child.text, edm_type, is_null)

        rows.append(row)

    next_url = None
    for link in root.findall(f"{{{ATOM_NS}}}link"):
        if link.attrib.get("rel") == "next":
            next_url = link.attrib.get("href")
            break

    return rows, next_url


def parse_json_payload(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], str | None]:
    if "value" in payload:
        return (
            payload.get("value", []),
            payload.get("@odata.nextLink") or payload.get("odata.nextLink"),
        )

    d = payload.get("d", {})

    if isinstance(d, dict):
        return d.get("results", []), d.get("__next")

    return [], None


def request_sharepoint_page(
    session: requests.Session,
    url: str,
    params: dict[str, str] | None = None,
    max_retries: int = 4,
) -> tuple[list[dict[str, Any]], str | None]:
    for attempt in range(1, max_retries + 1):
        response = session.get(url, params=params, timeout=90)

        if response.status_code in {429, 500, 502, 503, 504} and attempt < max_retries:
            wait_seconds = min(2**attempt, 30)
            logging.warning(
                "HTTP %s. Reintento %s/%s en %ss.",
                response.status_code,
                attempt,
                max_retries,
                wait_seconds,
            )
            time.sleep(wait_seconds)
            continue

        logging.info("GET %s", response.url)
        logging.info(
            "HTTP %s | Content-Type: %s",
            response.status_code,
            response.headers.get("Content-Type"),
        )

        if response.status_code != 200:
            logging.error(response.text[:2000])
            response.raise_for_status()

        content_type = response.headers.get("Content-Type", "").lower()

        if "json" in content_type:
            return parse_json_payload(response.json())

        if "xml" in content_type or "atom" in content_type:
            return parse_atom_feed(response.text)

        try:
            return parse_json_payload(response.json())
        except Exception:
            return parse_atom_feed(response.text)

    raise RuntimeError(f"No se pudo completar la petición: {url}")


def get_list_fields(session: requests.Session, settings: Settings) -> list[dict[str, Any]]:
    escaped_list_name = settings.source_list_name.replace("'", "''")

    url = f"{settings.source_site_url}/_api/web/lists/GetByTitle('{escaped_list_name}')/fields"

    params = {
        "$select": "InternalName,Title,Hidden,ReadOnlyField,FromBaseType,TypeAsString",
        "$top": "5000",
    }

    fields: list[dict[str, Any]] = []

    while url:
        batch, next_url = request_sharepoint_page(session, url, params=params)
        fields.extend(batch)

        url = next_url
        params = None

    logging.info("Campos encontrados en la lista: %s", len(fields))
    return fields


def get_clean_select_fields(session: requests.Session, settings: Settings) -> list[str]:
    fields = get_list_fields(session, settings)
    exclude_lower = settings.exclude_columns_lower

    # Campos base que sí suelen ser útiles y seguros
    allowed_base_fields = {
        "ID",
        "Title",
        "Created",
        "Modified",
    }

    selected: list[str] = []
    skipped: list[str] = []

    for field in fields:
        internal_name = str(field.get("InternalName", "")).strip()
        type_as_string = str(field.get("TypeAsString", "")).strip()

        if not internal_name:
            continue

        hidden = parse_sp_bool(field.get("Hidden"))
        from_base_type = parse_sp_bool(field.get("FromBaseType"))

        if internal_name.lower() in exclude_lower:
            skipped.append(f"{internal_name} excluded")
            continue

        if hidden:
            skipped.append(f"{internal_name} hidden")
            continue

        if type_as_string in UNSAFE_SELECT_TYPES:
            skipped.append(f"{internal_name} type={type_as_string}")
            continue

        # Evita traer basura base de SharePoint.
        # Se permiten solo Title, Created y Modified.
        if from_base_type and internal_name not in allowed_base_fields:
            skipped.append(f"{internal_name} base")
            continue

        selected.append(internal_name)

    selected = list(dict.fromkeys(selected))

    if not selected:
        logging.warning(
            "No se han podido seleccionar campos automáticamente. "
            "Se leerán todos los campos y se eliminarán columnas excluidas después."
        )
        return []

    logging.info("Campos seleccionados para $select: %s", ", ".join(selected))
    logging.info("Columnas excluidas por configuración: %s", ", ".join(settings.exclude_columns))

    if skipped:
        logging.info("Campos descartados automáticamente: %s", ", ".join(skipped[:80]))
        if len(skipped) > 80:
            logging.info("Campos descartados restantes: %s", len(skipped) - 80)

    return selected


def read_sharepoint_list(settings: Settings) -> list[dict[str, Any]]:
    session = make_session()

    escaped_list_name = settings.source_list_name.replace("'", "''")
    url = f"{settings.source_site_url}/_api/web/lists/GetByTitle('{escaped_list_name}')/items"

    params = {
        "$top": str(settings.page_size),
    }

    if settings.columns_mode == "auto":
        select_fields = get_clean_select_fields(session, settings)
        if select_fields:
            params["$select"] = ",".join(select_fields)

    rows: list[dict[str, Any]] = []

    while url:
        batch, next_url = request_sharepoint_page(session, url, params=params)

        rows.extend(batch)
        logging.info("Filas acumuladas: %s", len(rows))

        url = next_url
        params = None

    return rows


def json_safe(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def normalize_dataframe(records: list[dict[str, Any]], settings: Settings) -> pd.DataFrame:
    df = pd.DataFrame(records)

    if df.empty:
        return df

    # Deduplicar columnas case-insensitively para evitar conflictos como ID vs Id
    seen = set()
    cols_to_keep = []
    for col in df.columns:
        col_lower = col.lower()
        if col_lower not in seen:
            seen.add(col_lower)
            cols_to_keep.append(col)
        else:
            logging.info("Eliminando columna duplicada (case-insensitive): %s", col)
    df = df[cols_to_keep].copy()

    columns_to_drop: list[str] = []

    if settings.drop_api_metadata:
        columns_to_drop.append("__metadata")

    exclude_lower = settings.exclude_columns_lower

    for column in df.columns:
        if column.lower() in exclude_lower:
            columns_to_drop.append(column)

    if columns_to_drop:
        df = df.drop(columns=list(dict.fromkeys(columns_to_drop)), errors="ignore")

    for column in df.columns:
        df[column] = df[column].map(json_safe)

    # Rename columns to match required_columns in models config
    df = df.rename(
        columns={
            "Title": "Titulo",
            "HoraInicio": "Hora Inicio",
            "HoraFin": "Hora Fin",
            "Ubicaci_x00f3_n": "Ubicacion",
            "UrlDetalleEvento": "Url Detalle Evento",
        }
    )

    return df


def dataframe_hash(df: pd.DataFrame) -> str:
    payload = df.to_csv(index=False).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue

        width = min(max(len(value) for value in values) + 2, 80)
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = width


def add_excel_table(worksheet, table_name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return

    end_col = get_column_letter(worksheet.max_column)
    table_ref = f"A1:{end_col}{worksheet.max_row}"

    table = Table(
        displayName=clean_excel_table_name(table_name),
        ref=table_ref,
    )

    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(table)


def write_excel(df: pd.DataFrame, settings: Settings) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Preservar hojas existentes en el Excel, si el archivo existe
    other_sheets = {}
    if settings.output_path.exists():
        try:
            xl = pd.ExcelFile(settings.output_path)
            for sheet in xl.sheet_names:
                if sheet != settings.sheet_name and sheet != "_metadata":
                    other_sheets[sheet] = xl.parse(sheet)
                    logging.info("Preservando hoja existente: %s", sheet)
        except Exception as e:
            logging.warning("No se pudo leer el archivo existente para preservar hojas: %s", e)

    with pd.ExcelWriter(
        settings.output_path,
        engine="openpyxl",
        datetime_format="yyyy-mm-dd hh:mm:ss",
    ) as writer:
        df.to_excel(writer, sheet_name=settings.sheet_name, index=False)

        workbook = writer.book
        worksheet = workbook[settings.sheet_name]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        add_excel_table(worksheet, settings.output_name)
        autofit_columns(worksheet)

        # Volver a escribir las otras hojas preservadas
        for sheet_name, sheet_df in other_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            autofit_columns(workbook[sheet_name])

        if settings.include_metadata_sheet:
            metadata = pd.DataFrame(
                [
                    {"clave": "generated_at", "valor": datetime.now().isoformat(timespec="seconds")},
                    {"clave": "source_site_url", "valor": settings.source_site_url},
                    {"clave": "source_list_name", "valor": settings.source_list_name},
                    {"clave": "output_name", "valor": settings.output_name},
                    {"clave": "columns_mode", "valor": settings.columns_mode},
                    {"clave": "excluded_columns", "valor": ",".join(settings.exclude_columns)},
                    {"clave": "rows", "valor": len(df)},
                    {"clave": "columns", "valor": len(df.columns)},
                    {"clave": "data_sha256", "valor": dataframe_hash(df)},
                ]
            )

            metadata.to_excel(writer, sheet_name="_metadata", index=False)
            autofit_columns(workbook["_metadata"])

    logging.info("Excel generado: %s", settings.output_path)
    return settings.output_path


# upload_excel function removed


def main() -> None:
    settings = Settings.from_env()
    configure_logging(settings)

    logging.info("Inicio exportación")
    logging.info("Origen: %s | Lista: %s", settings.source_site_url, settings.source_list_name)
    logging.info("Excel local: %s", settings.output_path)
    logging.info("Destino: %s", settings.dest_relative_file_url)
    logging.info("COLUMNS_MODE: %s", settings.columns_mode)
    logging.info("EXCLUDE_COLUMNS: %s", ", ".join(settings.exclude_columns))

    records = read_sharepoint_list(settings)
    logging.info("Total registros leídos: %s", len(records))

    df = normalize_dataframe(records, settings)
    logging.info("DataFrame final: %s filas, %s columnas", len(df), len(df.columns))
    logging.info("Columnas finales: %s", ", ".join(df.columns))

    if settings.prevent_empty_upload and df.empty:
        raise RuntimeError("La lista se ha leído vacía. Se cancela la subida para no sobrescribir el archivo destino.")

    output_path = write_excel(df, settings)
    logging.info("Exportación local finalizada con éxito en: %s", output_path)


if __name__ == "__main__":
    main()
